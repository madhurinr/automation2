import openpyxl
book=openpyxl.load_workbook("D:testdatasheet.xlsx")
# print(book.active.title)
ss=book['userdetails']
# print(ss['A1'].value)
# print(ss.cell(1,1).value)
def colvalue():
    rown=ss.max_row
    col=ss.max_column
    for r in range(2,rown+1):
        for c in range(1,col+1):
            val= ss.cell(r,c).value
            print(val)

def getcellval(row,col):
    print(row)
    print(col)

    return ss.cell(row,col).value

def readmax_row(sheetname):
    ss=book[sheetname]
    return ss.max_row

def readmaxcol(sheetname):
    ss=book[sheetname]

def returncell(sheetname,row,col):
    print("Printing")
    print(sheetname)
    print(row)
    print(col)

    ss=book[sheetname]
    return ss.cell(row,col).value

print(readmax_row('userdetails'))

print(returncell('userdetails',1,1))

colvalue()