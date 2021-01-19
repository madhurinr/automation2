import openpyxl
#load worksheet
wb=openpyxl.load_workbook("D:/testxl.xlsx")
print(wb.active)
print(wb.sheetnames)
#create an object for any worksheet u want to work
ss=wb['SheetA']
print(ss.title)
#read data from cell using sheet object
c1=ss["A1"]
print(c1.value)
#ready data from cell using cell object
c2=ss.cell(2,2)
print(c2.value)
ct=ss.cell(3,2)
print(ct.value)
print(ss["B3"].value)
row=ss.max_row
col=ss.max_column
for i in range (1,row+1):
    for j in range(1,col+1):
        val=ss.cell(i,j)
        print(val.value)
for h in ss['A1':'C3']:
    for c in h:
        print(c.value)
