import openpyxl

wb= openpyxl.Workbook()
sh=wb.active
sh.title="testingworld"
sh['A1'].value="madhuri"



wb.create_sheet(title='nextsheet')

wb['nextsheet']
wb['nextsheet']['A3'].value=123
wb['nextsheet'].cell(2,3).value=223
wb.save("D:/newexcel.xlsx")