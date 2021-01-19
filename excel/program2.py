import openpyxl
wk=openpyxl.load_workbook("D:/testxl.xlsx")
print(wk.sheetnames)
print(wk.active)
sh=wk["SheetB"]
print(sh.title)
sh2=sh["A1"]
print(sh2.value)
cc=sh.cell(2,3)
print(cc.value)
row=sh.max_row
col=sh.max_column
for i in range(1,row+1):
    for j in range(1,col+1):
        val=sh.cell(i,j)
        print(val.value)
print("second way")
for h in sh['A1':'C3']:
    for j in h:
        print(j.value)