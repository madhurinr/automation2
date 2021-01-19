import openpyxl
wb=openpyxl.Workbook()
ss=wb.active
ss.title="father"
ss['A1']="revanna"
sm=wb.create_sheet('mother')
sm['A2']="Annapurna"
sss=wb.create_sheet('sister')
sss.cell(2,1).value='viji'
sss.cell(1,3).value='vini'
sh3=wb.create_sheet("del")
sh3.cell(2,2).value=32432
wb.remove(sh3)
wb.save('D:/testexcel.xlsx')
openw=openpyxl.load_workbook("D:/testexcel.xlsx")
print(openw.sheetnames)
father=openw['father']
print(father['A1'].value)
mother=openw['mother']
print(mother['A2'].value)
sister=openw['sister']
print(sister.cell(1,3).value)



